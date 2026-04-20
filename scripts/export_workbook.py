from __future__ import annotations

import json
import re
import sys
import unicodedata
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
BUNDLED_PORTFOLIO_PATH = ROOT / "data" / "portfolio-sample.json"
LOCAL_PORTFOLIO_PATH = ROOT / "data" / "portfolio.local.json"
DEFAULT_WORKBOOK_FILENAME: str | None = None
ASSET_METADATA_DEFAULTS = {
    "XRP": {
        "name": "XRP",
        "symbol": "KRW-XRP",
        "market": "crypto",
        "currency": "KRW",
        "priceSource": "upbit",
    },
    "ETH": {
        "name": "ETH",
        "symbol": "KRW-ETH",
        "market": "crypto",
        "currency": "KRW",
        "priceSource": "upbit",
    },
}


def normalize(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def parse_currency_cell(value: Any) -> int | float | None:
    normalized = normalize(value)
    if normalized is None:
        return None
    if isinstance(normalized, (int, float)):
        return normalize(normalized)
    if not isinstance(normalized, str):
        return None

    compact = normalized.replace(",", "").replace("원", "").strip()
    if not compact:
        return None
    if re.fullmatch(r"-?\d+", compact):
        return int(compact)
    if re.fullmatch(r"-?\d+(?:\.\d+)?", compact):
        return normalize(float(compact))
    return None


def list_workbooks() -> list[Path]:
    return sorted(
        [
            path
            for path in ROOT.glob("*.xlsx")
            if path.is_file() and not path.name.startswith("~$")
        ],
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )


def workbook_path(preferred_name: str | None = None) -> Path:
    workbooks = list_workbooks()
    if not workbooks:
        raise FileNotFoundError("No workbook was found in the project root.")

    target_name = preferred_name or DEFAULT_WORKBOOK_FILENAME
    if target_name:
        normalized_target_name = unicodedata.normalize("NFC", target_name)
        for path in workbooks:
            if unicodedata.normalize("NFC", path.name) == normalized_target_name:
                return path

    return workbooks[0]


def load_existing_portfolio() -> dict[str, Any]:
    for current_path in (LOCAL_PORTFOLIO_PATH, BUNDLED_PORTFOLIO_PATH):
        if not current_path.exists():
            continue
        try:
            return json.loads(current_path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            continue
    return {}


def resolve_strategy(existing: dict[str, Any], parsed_strategy: dict[str, Any]) -> dict[str, Any]:
    existing_strategy = existing.get("strategy")
    if isinstance(existing_strategy, dict) and {"selection", "entry", "exit", "stops"} <= set(existing_strategy):
        return existing_strategy
    return parsed_strategy


def enrich_holdings(existing: dict[str, Any], parsed_holdings: list[dict[str, Any]]) -> list[dict[str, Any]]:
    metadata_keys = {"name", "symbol", "market", "currency", "priceSource"}
    existing_holding_meta = {}

    for item in existing.get("holdings", []):
        if not isinstance(item, dict):
            continue
        key = (item.get("platform"), item.get("asset"))
        existing_holding_meta[key] = {
            meta_key: item.get(meta_key) for meta_key in metadata_keys if item.get(meta_key) is not None
        }

    enriched_holdings = []
    for item in parsed_holdings:
        key = (item.get("platform"), item.get("asset"))
        metadata = {
            **ASSET_METADATA_DEFAULTS.get(item.get("asset"), {}),
            **existing_holding_meta.get(key, {}),
        }
        enriched_holdings.append(
            {
                **item,
                **metadata,
                "name": metadata.get("name") or item.get("asset"),
            }
        )

    return enriched_holdings


def basis_year(label: str | None) -> int:
    if not label:
        return date.today().year
    match = re.search(r"(\d{4})", label)
    return int(match.group(1)) if match else date.today().year


def parse_trade_date(value: str | None, year: int) -> tuple[int, int, int]:
    if not value:
        return (year, 1, 1)
    month, day = (value.split("/") + ["1", "1"])[:2]
    return (year, int(month), int(day))


def parse_realized_trade_label(label: str | None) -> tuple[str | None, int | None]:
    normalized = normalize(label)
    if not normalized:
        return (None, None)

    match = re.match(r"^(?P<asset>.+?)\s+(?P<quantity>\d+)주\s+매도$", normalized)
    if not match:
        return (normalized, None)

    return (normalize(match.group("asset")), int(match.group("quantity")))


def is_section_header(value: Any) -> bool:
    return isinstance(value, str) and normalize(value).startswith("▎")


def find_row(sheet, column: str, expected: str) -> int:
    normalized_expected = unicodedata.normalize("NFC", expected)
    for row in range(1, sheet.max_row + 1):
        value = normalize(sheet[f"{column}{row}"].value)
        if not isinstance(value, str):
            continue
        if unicodedata.normalize("NFC", value) == normalized_expected:
            return row
    raise ValueError(f"Could not find '{expected}' in {sheet.title}!{column}:{column}")


def parse_asset_status(sheet) -> list[dict[str, Any]]:
    section_row = find_row(sheet, "A", "▎ 자산별 현황")
    asset_status = []

    for row in range(section_row + 2, sheet.max_row + 1):
        category = normalize(sheet[f"A{row}"].value)
        if category == "현금" or category == "총 자산" or is_section_header(category):
            break
        if not category:
            continue

        asset_status.append(
            {
                "category": category,
                "platform": normalize(sheet[f"B{row}"].value),
                "valuation": normalize(sheet[f"C{row}"].value) or 0,
                "principal": normalize(sheet[f"D{row}"].value) or 0,
                "pnl": normalize(sheet[f"E{row}"].value) or 0,
                "returnRate": normalize(sheet[f"F{row}"].value) or 0,
            }
        )

    return asset_status


def parse_cash_positions(sheet) -> list[dict[str, Any]]:
    section_row = find_row(sheet, "A", "▎ 자산별 현황")
    cash_positions = []

    for row in range(section_row + 2, sheet.max_row + 1):
        category = normalize(sheet[f"A{row}"].value)
        if category == "총 자산" or is_section_header(category):
            break
        if category not in ("현금", None):
            continue

        platform = normalize(sheet[f"B{row}"].value)
        if not platform:
            continue

        cash_positions.append(
            {
                "platform": platform,
                "amount": normalize(sheet[f"D{row}"].value) or 0,
            }
        )

    return cash_positions


def parse_holdings(sheet) -> list[dict[str, Any]]:
    section_row = find_row(sheet, "A", "▎ 보유종목 상세")
    holdings = []

    for row in range(section_row + 2, sheet.max_row + 1):
        platform = normalize(sheet[f"A{row}"].value)
        if is_section_header(platform):
            break
        if not platform:
            continue

        holdings.append(
            {
                "platform": platform,
                "asset": normalize(sheet[f"B{row}"].value),
                "quantity": normalize(sheet[f"C{row}"].value) or 0,
                "averagePrice": normalize(sheet[f"D{row}"].value) or 0,
                "valuation": normalize(sheet[f"E{row}"].value) or 0,
                "returnRate": normalize(sheet[f"F{row}"].value) or 0,
            }
        )

    return holdings


def parse_realized_summary(sheet) -> list[dict[str, Any]]:
    section_row = find_row(sheet, "A", "▎ 국내주식 실현손익 요약")
    realized = []
    for row in range(section_row + 2, sheet.max_row + 1):
        platform = normalize(sheet[f"A{row}"].value)
        if not platform:
            continue
        if isinstance(platform, str) and platform.startswith("실현손익 합계"):
            break

        label = normalize(sheet[f"B{row}"].value)
        if not label:
            continue

        asset_name, quantity = parse_realized_trade_label(label)
        realized.append(
            {
                "platform": platform,
                "asset": label,
                "assetName": asset_name,
                "quantity": quantity,
                "pnl": normalize(sheet[f"C{row}"].value) or 0,
                "returnRate": normalize(sheet[f"D{row}"].value) or 0,
            }
        )

    return realized


def parse_stock_trades(sheet) -> list[dict[str, Any]]:
    stock_trades = []

    for broker in ("카카오증권", "미래에셋"):
        section_row = find_row(sheet, "A", f"▎ {broker} 매매 내역")
        for row in range(section_row + 2, sheet.max_row + 1):
            date_value = normalize(sheet[f"A{row}"].value)
            if is_section_header(date_value):
                break
            if not date_value:
                continue

            stock_trades.append(
                {
                    "date": date_value,
                    "broker": broker,
                    "asset": normalize(sheet[f"B{row}"].value),
                    "side": normalize(sheet[f"C{row}"].value),
                    "quantity": normalize(sheet[f"D{row}"].value) or 0,
                    "price": normalize(sheet[f"E{row}"].value) or 0,
                    "amount": normalize(sheet[f"F{row}"].value) or 0,
                    "fee": normalize(sheet[f"G{row}"].value) or 0,
                    "note": normalize(sheet[f"H{row}"].value),
                }
            )

    return stock_trades


def parse_crypto_trades(sheet) -> list[dict[str, Any]]:
    section_row = find_row(sheet, "A", "▎ 매매 내역")
    crypto_trades = []

    for row in range(section_row + 2, sheet.max_row + 1):
        date_value = normalize(sheet[f"A{row}"].value)
        if is_section_header(date_value):
            break
        if not date_value:
            continue

        crypto_trades.append(
            {
                "date": date_value,
                "asset": normalize(sheet[f"G{row}"].value),
                "side": normalize(sheet[f"B{row}"].value),
                "quantity": normalize(sheet[f"C{row}"].value) or 0,
                "price": normalize(sheet[f"D{row}"].value) or 0,
                "amount": normalize(sheet[f"E{row}"].value) or 0,
                "fee": normalize(sheet[f"F{row}"].value) or 0,
            }
        )

    return crypto_trades


def pop_matching_sell_trade(
    stock_sells: list[dict[str, Any]], realized_item: dict[str, Any]
) -> dict[str, Any] | None:
    match_index = next(
        (
            index
            for index, trade in enumerate(stock_sells)
            if trade["broker"] == realized_item["platform"]
            and trade["asset"] == realized_item["assetName"]
            and trade["quantity"] == realized_item["quantity"]
        ),
        None,
    )
    if match_index is None:
        match_index = next(
            (
                index
                for index, trade in enumerate(stock_sells)
                if trade["broker"] == realized_item["platform"]
                and trade["asset"] == realized_item["assetName"]
            ),
            None,
        )
    if match_index is None:
        match_index = next(
            (
                index
                for index, trade in enumerate(stock_sells)
                if trade["asset"] == realized_item["assetName"]
                and trade["quantity"] == realized_item["quantity"]
            ),
            None,
        )

    return stock_sells.pop(match_index) if match_index is not None else None


def principal_from_realized(item: dict[str, Any]) -> float:
    rate = item.get("returnRate") or 0
    pnl = item.get("pnl") or 0
    return (pnl / rate) if rate else 0.0


def group_realized_trades(
    realized: list[dict[str, Any]], stock_trades: list[dict[str, Any]]
) -> list[dict[str, Any]]:
    stock_sells = [trade.copy() for trade in stock_trades if trade["side"] == "매도"]
    grouped: list[dict[str, Any]] = []
    grouped_index: dict[tuple[Any, ...], int] = {}

    for item in realized:
        matched_trade = pop_matching_sell_trade(stock_sells, item)
        trade_date = matched_trade["date"] if matched_trade else None
        asset_name = item["assetName"] or item["asset"]
        group_key = (item["platform"], asset_name, trade_date)
        principal = principal_from_realized(item)

        if group_key not in grouped_index:
            grouped_index[group_key] = len(grouped)
            grouped.append(
                {
                    "platform": item["platform"],
                    "asset": item["asset"],
                    "assetName": asset_name,
                    "quantity": item["quantity"] or (matched_trade["quantity"] if matched_trade else None),
                    "pnl": item["pnl"],
                    "returnRate": item["returnRate"],
                    "date": trade_date,
                    "_principal": principal,
                }
            )
            continue

        grouped_item = grouped[grouped_index[group_key]]
        grouped_item["quantity"] = (grouped_item["quantity"] or 0) + (
            item["quantity"] or (matched_trade["quantity"] if matched_trade else 0)
        )
        grouped_item["pnl"] += item["pnl"]
        grouped_item["_principal"] += principal

    for item in grouped:
        quantity = item["quantity"]
        if item["_principal"]:
            item["returnRate"] = item["pnl"] / item["_principal"]
        if quantity:
            quantity_label = int(quantity) if isinstance(quantity, float) and quantity.is_integer() else quantity
            item["asset"] = f"{item['assetName']} {quantity_label}주 매도"
        del item["_principal"]

    return grouped


def build_chart_data(
    *,
    basis_date_label: str | None,
    holdings: list[dict[str, Any]],
    realized: list[dict[str, Any]],
) -> dict[str, Any]:
    year = basis_year(basis_date_label)
    returns_by_asset: dict[str, dict[str, float]] = {}
    for item in holdings:
        if not item["quantity"] or not item["valuation"]:
            continue
        asset = item["asset"]
        principal = (
            item["valuation"] / (1 + item["returnRate"])
            if (1 + item["returnRate"]) not in (0, 0.0)
            else item["averagePrice"] * item["quantity"]
        )
        bucket = returns_by_asset.setdefault(
            asset,
            {
                "label": asset,
                "valuation": 0.0,
                "principal": 0.0,
            },
        )
        bucket["valuation"] += item["valuation"]
        bucket["principal"] += principal

    returns_comparison = []
    for asset, values in returns_by_asset.items():
        pnl = values["valuation"] - values["principal"]
        principal = values["principal"] or 1
        returns_comparison.append(
            {
                "label": asset,
                "valuation": round(values["valuation"]),
                "pnl": round(pnl),
                "returnRate": pnl / principal,
            }
        )
    returns_comparison.sort(key=lambda item: item["returnRate"], reverse=True)

    realized_trade_points = []
    for order, item in enumerate(realized):
        trade_date = item.get("date")
        asset_name = item.get("assetName") or item["asset"]

        realized_trade_points.append(
            {
                "date": trade_date,
                "asset": asset_name,
                "broker": item["platform"],
                "quantity": item["quantity"] or 0,
                "pnl": item["pnl"],
                "sortKey": parse_trade_date(trade_date, year),
                "order": order,
            }
        )

    realized_trade_points.sort(key=lambda item: (item["sortKey"], item["order"]))

    realized_history = []
    cumulative = 0
    for trade in realized_trade_points:
        if realized_history and realized_history[-1]["date"] == trade["date"]:
            realized_history[-1]["dailyPnl"] += trade["pnl"]
            realized_history[-1]["tradeCount"] += 1
            realized_history[-1]["items"].append(
                f"{trade['asset']} {trade['quantity']}주 · {trade['broker']}"
            )
            cumulative += trade["pnl"]
            realized_history[-1]["cumulativePnl"] = cumulative
            continue

        cumulative += trade["pnl"]
        realized_history.append(
            {
                "date": trade["date"],
                "displayDate": f"{trade['sortKey'][0]}-{trade['sortKey'][1]:02d}-{trade['sortKey'][2]:02d}",
                "dailyPnl": trade["pnl"],
                "cumulativePnl": cumulative,
                "tradeCount": 1,
                "items": [f"{trade['asset']} {trade['quantity']}주 · {trade['broker']}"],
            }
        )

    return {
        "returnsComparison": returns_comparison,
        "realizedHistory": realized_history,
    }


def parse_strategy(sheet) -> dict[str, Any]:
    entry_steps = []
    for row in range(8, 11):
        entry_steps.append(
            {
                "label": normalize(sheet[f"A{row}"].value),
                "allocation": normalize(sheet[f"B{row}"].value),
                "trigger": normalize(sheet[f"C{row}"].value),
                "splitGuide": normalize(sheet[f"D{row}"].value),
            }
        )

    exit_rules = []
    for row in range(17, 21):
        exit_rules.append(
            {
                "label": normalize(sheet[f"A{row}"].value),
                "trigger": normalize(sheet[f"B{row}"].value),
                "action": normalize(sheet[f"C{row}"].value),
                "note": normalize(sheet[f"D{row}"].value),
            }
        )

    checklist = []
    for row in range(24, 31):
        value = normalize(sheet[f"A{row}"].value)
        if value:
            checklist.append(value)

    return {
        "title": normalize(sheet["A2"].value),
        "entryPrinciple": normalize(sheet["B6"].value),
        "entryNotes": [normalize(sheet["A11"].value), normalize(sheet["A12"].value)],
        "entrySteps": entry_steps,
        "exitPrinciple": normalize(sheet["B15"].value),
        "exitRules": exit_rules,
        "checklistTitle": normalize(sheet["A22"].value),
        "checklistWindow": normalize(sheet["A23"].value),
        "checklist": checklist,
    }


def parse_xrp_defense(sheet) -> dict[str, Any]:
    row_mapping = {
        8: "initialQuantity",
        9: "initialAveragePrice",
        10: "soldQuantity",
        11: "averageSellNet",
        12: "rebuyQuantity",
        13: "averageRebuyGross",
        14: "defenseGain",
        15: "finalAveragePrice",
        16: "averageCutAmount",
        17: "averageCutRate",
        18: "realizedPnl",
        19: "remainingQuantity",
        20: "breakevenTargetBuyPrice",
    }
    return {key: normalize(sheet[f"J{row}"].value) for row, key in row_mapping.items()}


def main() -> None:
    preferred_workbook = sys.argv[1] if len(sys.argv) > 1 else None
    source = workbook_path(preferred_workbook)
    workbook_values = load_workbook(source, data_only=True)
    existing_portfolio = load_existing_portfolio()

    overview = workbook_values["총괄현황"]
    stocks = workbook_values["국내주식 매매일지"]
    crypto = workbook_values["업비트 매매일지"]
    strategy_sheet = workbook_values["페이즈1 전략"]
    calc = workbook_values["_calc"]

    asset_status = parse_asset_status(overview)
    cash_positions = parse_cash_positions(overview)
    holdings = enrich_holdings(existing_portfolio, parse_holdings(overview))
    stock_trades = parse_stock_trades(stocks)
    crypto_trades = parse_crypto_trades(crypto)
    realized = group_realized_trades(parse_realized_summary(overview), stock_trades)

    initial_investment = parse_currency_cell(crypto["F4"].value) or 0
    asset_valuation_total = sum(item["valuation"] for item in asset_status)
    invested_principal = sum(item["principal"] for item in asset_status)
    portfolio_pnl = sum(item["pnl"] for item in asset_status)
    cash_total = sum(item["amount"] for item in cash_positions)
    realized_total = sum(item["pnl"] for item in realized)
    chart_data = build_chart_data(
        basis_date_label=normalize(overview["A3"].value),
        holdings=holdings,
        realized=realized,
    )

    portfolio = {
        "metadata": {
            "title": normalize(overview["A2"].value),
            "mantra": normalize(overview["A1"].value),
            "basisDateLabel": normalize(overview["A3"].value),
            "workbook": source.name,
        },
        "summary": {
            "initialInvestment": initial_investment,
            "totalAssets": normalize(overview["D13"].value) or 0,
            "investedPrincipal": invested_principal,
            "assetValuationTotal": asset_valuation_total,
            "cashTotal": cash_total,
            "portfolioPnl": portfolio_pnl,
            "portfolioReturnRate": (portfolio_pnl / invested_principal) if invested_principal else 0,
            "realizedProfitTotal": realized_total,
            "liquidityRatio": (cash_total / (normalize(overview["D13"].value) or 1)),
        },
        "assetStatus": asset_status,
        "cashPositions": cash_positions,
        "holdings": holdings,
        "realized": realized,
        "charts": chart_data,
        "strategy": resolve_strategy(existing_portfolio, parse_strategy(strategy_sheet)),
        "trades": {
            "stocks": stock_trades,
            "crypto": crypto_trades,
        },
        "analytics": {
            "prices": {
                "samsungElectronics": normalize(calc["I1"].value),
                "skHynix": normalize(calc["I2"].value),
                "xrp": normalize(calc["I3"].value),
                "eth": normalize(calc["I7"].value),
            },
            "xrpDefense": parse_xrp_defense(crypto),
        },
    }

    if isinstance(existing_portfolio.get("targets"), dict):
        portfolio["targets"] = existing_portfolio["targets"]
    if isinstance(existing_portfolio.get("strategyBudgets"), dict):
        portfolio["strategyBudgets"] = existing_portfolio["strategyBudgets"]

    payload = json.dumps(portfolio, ensure_ascii=False, indent=2)
    output_path = LOCAL_PORTFOLIO_PATH

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(payload, encoding="utf-8")
    print(f"Wrote {output_path.relative_to(ROOT)} from {source.name}")


if __name__ == "__main__":
    main()
